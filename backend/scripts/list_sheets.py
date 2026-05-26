"""Helper rapido para que el backend Express obtenga la lista de hojas de un
Excel sin cargar todos los datos (read_only=True). Se invoca como:

    python3 list_sheets.py <ruta_al_xlsx>

Imprime un JSON con la forma {"sheets": ["Licitaciones_lineas", "INPUT_710", ...]}
en stdout. Codigo de salida 0 = ok, 1 = error (mensaje en stderr).
"""
import json
import sys

try:
    from openpyxl import load_workbook
except ImportError as exc:
    print(f"openpyxl no esta instalado: {exc}", file=sys.stderr)
    sys.exit(1)


def main() -> int:
    if len(sys.argv) != 2:
        print("Uso: python3 list_sheets.py <archivo.xlsx>", file=sys.stderr)
        return 1

    path = sys.argv[1]
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception as exc:  # noqa: BLE001
        print(f"No se pudo abrir el Excel: {exc}", file=sys.stderr)
        return 1

    print(json.dumps({"sheets": sheets}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
